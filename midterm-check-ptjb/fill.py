import openpyxl

crs = int(input("請輸入總課程數:"))

# excel 1-2 from google form審查意見
gf = openpyxl.load_workbook('1-2.xlsx')
shtgf = gf.worksheets[0]

# excel 2-2 目標寫入對象
tg = openpyxl.load_workbook('2-2.xlsx')
shttg = tg.worksheets[0]

for v in range(crs):
    if shtgf.cell(row=v+2, column=3).value==None:
        break

    for lp in range(crs):
        if shttg.cell(row=5+lp*3, column=7).value==None: break
        if shtgf.cell(row=v+2, column=3).value[0:6] == shttg.cell(row=5+lp*3, column=7).value:
            idx=lp
            break

    # 經費執行情形
    for e in range(2):
        sum = 0
        for f in range(3):
            sum += int(shtgf.cell(row=v+2, column=4+f+e*3).value)
            shttg.cell(row=5+idx*3+e, column=9+f).value = format(int(shtgf.cell(row=v+2, column=4+f+e*3).value), ',')
        if e==0: 
            sum1=sum
        elif e==1:
            sum2=sum
        shttg.cell(row=5+idx*3+e, column=12).value = format(sum, ',')
    shttg.cell(row=5+idx*3+2, column=9).value = \
        '{:.2f}%'.format(100*sum2/sum1)

    # 教學設備採購進度
    shttg.cell(row=5+idx*3, column=13).value = shtgf.cell(row=v+2, column=10).value
    shttg.cell(row=5+idx*3, column=14).value = shtgf.cell(row=v+2, column=11).value

    # 模組課程結合情形
    mod = shtgf.cell(row=v+2, column=12).value.split(', ')
    hr = shtgf.cell(row=v+2, column=13).value.split('、')
    cnt = 0
    for m in mod:
        nm=''
        shttg.cell(row=5+idx*3+cnt, column=15).value = m.split()[0]
        for mnm in range(len(m.split())):
            if mnm!=0: nm = nm + m.split()[mnm] + ' '
        shttg.cell(row=5+idx*3+cnt, column=16).value = nm
        shttg.cell(row=5+idx*3+cnt, column=17).value = '使用教材 '+hr[cnt].split()[1][1:-1]
        cnt+=1
    
    # 課程開授成效
    for e in range(2):
        st=''
        st+='1. 修課人次:'
        st+=str(shtgf.cell(row=v+2, column=15+e*3).value)
        st+='人次\n'

        st+='2. 專題作品數:'
        st+=str(shtgf.cell(row=v+2, column=16+e*3).value)
        st+='件\n'

        st+='3. 質化成效說明:\n'
        st+=str(shtgf.cell(row=v+2, column=17+e*3).value)

        shttg.cell(row=5+idx*3, column=18+e).value = st

    # 參加聯盟活動競賽情形
    for e in range(2):
        st=''
        st+='1. 參與聯盟相關課程推廣研習、座談：'
        st+=str(shtgf.cell(row=v+2, column=21+e*4).value)
        st+='人次'
        st+=str(shtgf.cell(row=v+2, column=22+e*4).value)
        st+='場次\n'

        st+='2. 參與聯盟相關競賽學生人數：'
        st+=str(shtgf.cell(row=v+2, column=23+e*4).value)
        st+='人\n'

        if shtgf.cell(row=v+2, column=24+e*4).value!=None:
            st+=str(shtgf.cell(row=v+2, column=24+e*4).value)

        shttg.cell(row=5+idx*3, column=20+e).value = st
    
    # 業界和校外講師參與教學情形
    shttg.cell(row=5+idx*3, column=22).value = str(shtgf.cell(row=v+2, column=29).value)

tg.save('result.xlsx')
