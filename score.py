import openpyxl

# 分數及意見總表 (此次寫入目標)
sc = openpyxl.load_workbook('allScore.xlsx')
shtsc = sc.worksheets[2] # 2代表終端聯盟

# excel 1-2 from google form審查意見
gf = openpyxl.load_workbook('1-2.xlsx')
shtgf = gf.worksheets[0]

# 審查委員list
comite = ['洪士灝', '馬席彬', '黃世旭', '吳文慶'] 

# 委員1之綜合分數的格子位置
shtgf.cell(row=4,column=11).value 

